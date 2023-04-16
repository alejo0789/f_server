import openpyxl
import os.path
import re
import datetime

# Esta función toma como entrada una cadena de texto y la guarda en un archivo Excel
def guardar_en_excel(texto, numero):

    # Define the trimester 
    trimestre_actual = define_trimestre(datetime.datetime.now())
    fecha_actual = datetime.datetime.now()

# Convertir la fecha actual en una cadena de texto con formato dd/mm/yyyy
    fecha_formateada = fecha_actual.strftime('%d/%m/%Y')
    
    # Creates the name of the file with the person's number and trimester
    #name = numero + '_' + trimestre_actual + ".xlsx"

    mes_actual = fecha_actual.strftime('%m')
    anio_actual = fecha_actual.strftime('%Y')

    # Crea el nombre del archivo con el número de persona, mes y año actual
    name = f"{numero}_{mes_actual}_{anio_actual}.xlsx"



    # Creates the path of the file
    folder_path = os.path.join('src', 'files', numero)
    ruta_archivo = os.path.join(folder_path, name)
    ruta_archivo_madre=os.path.join('src', 'files', 'template_presupuesto.xlsx')
    # Creates the folder if it doesn't exist
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Load the existing Excel file or create a new one if it doesn't exist
    if os.path.isfile(ruta_archivo):
        # If the file exists, load the Excel workbook
        libro_excel = openpyxl.load_workbook(ruta_archivo)
        # Use the active sheet of the Excel workbook
        hoja = libro_excel['Dia a dia']
    else:
        # If the file doesn't exist, create a new one
        libro_excel = openpyxl.load_workbook(ruta_archivo_madre)
        hoja = libro_excel.active

    texto, valor = separar_texto_y_numeros(texto)

    tipo, texto = define_tipo(texto.lower())



    for fila in range(5, hoja.max_row + 1):
     if hoja.cell(row=fila, column=4).value is None:
        fila_vacia = fila
        break
    else:
     fila_vacia = hoja.max_row + 1
    hoja.cell(row=fila_vacia, column=2, value=fecha_formateada)
    hoja.cell(row=fila_vacia, column=3, value=texto)
    hoja.cell(row=fila_vacia, column=4, value=valor)
    hoja.cell(row=fila_vacia, column=5, value=tipo)

    # Save the Excel file to disk
    libro_excel.save(ruta_archivo)



# Esta función toma como entrada una cadena de texto y devuelve dos listas, una con las letras y otra con los números

def separar_texto_y_numeros(cadena):
    # Buscar grupos de letras y grupos de números en la cadena
    grupos = re.findall(r'[a-zA-Z]+|\d+', cadena)
    
    # Inicializar listas para texto y números
    texto = []
    numeros = []
    
    for grupo in grupos:
        # Intentar convertir el grupo a un número
        try:
            numero = float(grupo)
            numeros.append(numero)
        except ValueError:
            # Si el grupo no es un número, agregarlo a la lista de texto
            texto.append(grupo)
    
    # Unir la lista de texto en una sola cadena con espacios entre cada palabra
    texto = ' '.join(texto)
   
    return texto, numeros[0] if numeros else None



def define_trimestre(fecha):
    # Obtener el trimestre actual
    trimestre = (fecha.month - 1) // 3 + 1

    # Asignar el nombre del archivo según el trimestre actual
    if trimestre == 1:
        trimestreactual= 'ene-mar'
    elif trimestre == 2:
        trimestreactual = 'abr-jun'
    elif trimestre == 3:
       trimestreactual = 'jul-sep'
    elif trimestre == 4:
       trimestreactual = 'oct-dic'

    return trimestreactual

def define_tipo(texto):

  if "ingreso" in texto:
   texto = texto.replace("ingreso", "")
   return 'Ingreso', texto
  elif "ahorro" in texto:
   texto = texto.replace("ahorro", "")
   return 'Ahorro', texto
  else:
     return 'Gasto', texto
     
     
  
